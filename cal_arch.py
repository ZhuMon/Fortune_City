from openpyxl import load_workbook
from openpyxl import Workbook

class citizen():
    def __init__(self):
        # 收入 其他 食物 飲品 交通 消費 遊戲 居家 醫藥 3C
        self.value = {}
        self.name = ""
        for i in range(1, 11):
            self.value[i] = 0
        
        self.hired = False

    def put(self, index, value):
        self.value[index] = value

    def get(self, index):
        return self.value[index]

    def is_hired(self):
        return self.hired

    def hire(self, kind):
        self.hired = True
        self.work = kind

    def max_income_work(self):
        return max(self.value)

    def max_income_work_value(self):
        return self.get(self.max_income_work())
def get_income_table(ws):
    income_table = {}
    for i in range(1,9): #level
        income_table[i] = {}
        for j in range(2, 12): # 熟練度
            v = ws.cell(row=j, column=i+1).value
            income_table[i][j-1] = v
    return income_table

class building():
    def __init__(self, level, kind):
        self.level = level
        self.kind = kind
        self.employees = [] # [citizen(), citizen(), ... ]

    def hire(self, person):
        if len(self.employees) < self.level:
            self.employees.append(person)
            return True
        else:
            return False

    def fire(self, person):
        if person in self.employees:
            self.emplyees.remove(person)
            return True
        else:
            return False

def num2kind(num, level):
    text = ["", "收入","其他","食物","飲品","交通","消費","遊戲","居家","醫藥","3C"]
    return text[num]+str(level)


def get_building_table(ws):
    building_table = {} # {1:{1:1, 3:1}, 2:{}..}
    building_list = [] # store building()
    for t in range(1, 11): # building type
        building_table[t] = {}
        for lv in range(1, 9):
            v = ws.cell(row = lv+1, column=t+1).value
            if v != 0:
                building_table[t][lv] = v
                for i in range(v):
                    b = building(lv, t)
                    building_list.append(b)
    return building_table, building_list

def get_citizens(ws):
    citizens = {}
    for c in range(2, ws.max_row+1):
        tmp = citizen()
        for t in range(2, 12):
            v = ws.cell(row=c, column=t).value
            tmp.put(t-1, v)
        tmp.name = ws.cell(row=c, column=1).value
        citizens[ws.cell(row=c, column=1).value] = tmp
    return citizens

def sort_building(building_list, building_table, new_list = []):
    if len(building_list) == 0:
        return new_list
    every_type_max_level = [0 for i in range(10)] 
    for t, b in building_table.items():
        every_type_max_level[t-1] = 0
        max_tmp = 0
        for lv, num in b.items():
            if num > max_tmp:
                every_type_max_level[t-1] = lv

    max_level = max(every_type_max_level)
    max_type = every_type_max_level.index(max_level)+1

    max_type_list = [i for i in building_list 
            if (i.kind == max_type and i.level == max_level)]
    new_list.extend(max_type_list)
    # pop max_type_list
    building_list = [i for i in building_list if i not in max_type_list]
    building_table[max_type].pop(max_level)

    # print(building_table)
    
    new_list = sort_building(building_list, building_table, new_list)

    return new_list

def get_highest_citizen(citizens, kind):
    max_value = 0
    great_citizen = ""
    for name, citizen in citizens.items():
        if citizen.get(kind) > max_value:
            great_citizen = name
            max_value = citizen.get(kind)

    return citizens[great_citizen]

def main():
    file_name = "~/Desktop/fortune_city.xlsx"
    wb = load_workbook("./fortune_city.xlsx")
    ws1 = wb.get_sheet_by_name("工作表1")
    income_table = get_income_table(ws1)

    ws_building = wb.get_sheet_by_name("building")
    building_table, building_list = get_building_table(ws_building)

    # print(building_table)
    ws_citizen = wb.get_sheet_by_name("citizen")
    citizens = get_citizens(ws_citizen)


    total_income = 0

    # 擁有最多最高等級建築的種類開始
    new_building_list = sort_building(building_list, building_table)
    # print(new_list)

    for n in new_building_list:
        print(num2kind(n.kind, n.level))
        for lv in range(n.level):
            if len(citizens) == 0:
                break
            c = get_highest_citizen(citizens, n.kind)
            if n.hire(c) == False:
                print("error")
                break
            total_income += int(income_table[n.level][c.get(n.kind)])
            print(int(income_table[n.level][c.get(n.kind)]))
            citizens.pop(c.name)
            print("\t"+c.name)
        
        if len(citizens) == 0:
            break

    print(total_income)

if __name__ == "__main__":
    main()
